import pandas as pd
import re
import os
import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule, FormulaRule

# --- 1. Cleanup Old Files ---
cleanup_dir = '1.Datenbank/Andreas - Analyse/Analyse-Ergebnisse/'
files_to_delete = [
    'Akteure_Status_Tracking.xlsx',
    'Akteure_Status_Tracking_Neu.xlsx',
    'Interview_Partner_Tracking.xlsx',
    'Interview_Partner_Tracking_Final.xlsx',
    'Interview_Partners_50_Final.xlsx'
]

print('--- Cleaning up old files ---')
for f in files_to_delete:
    full_path = os.path.join(cleanup_dir, f)
    if os.path.exists(full_path):
        try:
            os.remove(full_path)
            print(f'Deleted: {f}')
        except Exception as e:
            print(f'Could not delete {f}: {e}')

# --- 2. Data Setup (The 52 Partners) ---
data = [
    # Cat 1
    ['Katholisches Familienzentrum St. Konrad', 'Soziales / Selbsthilfe', 'Villingen-Schwenningen', 'kontakt@familienzentrum-stkonrad.de', 'Groß & Regelmäßig'],
    ['Ago e.V.', 'Prävention', 'Donaueschingen', 'info@verein-ago.de', 'Groß & Regelmäßig'],
    ['Psychologische Beratungsstelle Ev./Kath. Kirche', 'Beratung', 'Villingen-Schwenningen', 'gemeindebuero.schwenningen@elk-wue.de', 'Groß & Regelmäßig'],
    ['Sommerrainschule St. Georgen -- Sportangebote', 'Sportangebot', 'St. Georgen', 'sbbz-st.georgen@t-online.de', 'Groß & Regelmäßig'],
    ['Kinderfeuerwehr St. Georgen', 'BOS', 'St. Georgen', 'info@feuerwehr-st-georgen.de', 'Groß & Regelmäßig'],
    ['Neckarschule -- Klasse 2000', 'Prävention', 'Villingen-Schwenningen', 'neckarschule@villingen-schwenningen.de', 'Groß & Regelmäßig'],
    ['Georg Müller Grund- und Realschule -- Präventionskonzept', 'Prävention', 'Villingen-Schwenningen', 'info@gmsvs.de', 'Groß & Regelmäßig'],
    ['Caritas Schwangerschaftsberatung', 'Beratung', 'Villingen-Schwenningen', 'info@caritas-sbk.de', 'Groß & Regelmäßig'],
    ['Sozialpsychiatrischer Dienst (Caritas)', 'Soziales', 'Villingen-Schwenningen', 'elke.schwarz@caritas-sbk.de', 'Groß & Regelmäßig'],
    ['Malteser Hilfsdienst -- Psychosoziale Notfallvorsorge', 'BOS', 'Villingen-Schwenningen', 'Rainer.Kuehl@malteser.org', 'Groß & Regelmäßig'],
    ['AWO Haus der Kinder', 'Betreuung', 'Villingen-Schwenningen', 'hausderkinder@awo-ov-vs.de', 'Groß & Regelmäßig'],
    ['Pro Familia Villingen-Schwenningen', 'Beratung', 'Villingen-Schwenningen', 'vs-villingen@profamilia.de', 'Groß & Regelmäßig'],
    ['Sportangebot Öfingen 1969', 'Sportangebot', 'Bad Dürrheim', 'jugend@sv-oefingen.de', 'Groß & Regelmäßig'],
    # Cat 2
    ['BEKJ Villingen-Schwenningen', 'Beratung', 'Villingen-Schwenningen', 'beratungsstelle-bekj-fw@Lrasbk.de', 'Groß & Gelegentlich'],
    ['Polizei Triberg -- Präventionsvorträge', 'Prävention', 'Triberg', 'Keine E-Mail!', 'Groß & Gelegentlich'],
    ['Schwarzwald-Gymnasium', 'Beratung', 'Triberg', 'rf@schwarzwald-gymnasium.de', 'Groß & Gelegentlich'],
    ['Lernort Natur -- Jägervereinigung SBK', 'Natur', 'SBK', 'kjm-kjv-sbk@outlook.de', 'Groß & Gelegentlich'],
    ['Naturfreunde', 'Umweltschutz', 'Bräunlingen', 'naturfreunde-braeunlingen@gmx.de', 'Groß & Gelegentlich'],
    ['Kontaktstelle Beratung & Freizeit', 'Soziales', 'Donaueschingen', 'David.Mueller@Bruderhausdiakonie.de', 'Groß & Gelegentlich'],
    ['Schulnetzwerk', 'Bildung', 'St. Georgen', 'sekretariat@rsstg.de', 'Groß & Gelegentlich'],
    ['AWO Kinder- und Jugendhilfe', 'Soziales', 'Villingen-Schwenningen', 's.brunner@awo-ov-vs.de', 'Groß & Gelegentlich'],
    ['DRK -- Psychosoziale Notfallvorsorge', 'BOS', 'Villingen-Schwenningen', 'mail@psnv-sbk.de', 'Groß & Gelegentlich'],
    ['Projekt Bewegte Schule -- Dom-Clemente-Schule', 'Sportangebot', 'Schonach', 'd.tritschler@schonach.de', 'Groß & Gelegentlich'],
    ['Schönwälder Naturfreibad', 'Sportstätte', 'Schönwald', 'mail@schoenwald.de', 'Groß & Gelegentlich'],
    ['Freibad Tannheim', 'Sportstätte', 'Tannheim', 'info@freibad-tannheim.de', 'Groß & Gelegentlich'],
    ['Deutscher Kinderschutzbund OV', 'Soziales', 'Furtwangen', 'info@dksb-furtwangen.de', 'Groß & Gelegentlich'],
    ['BUND Ortsverband', 'Umweltschutz', 'Königsfeld', 'bund.sbh@bund.net', 'Groß & Gelegentlich'],
    # Cat 3
    ['Bundesverband Psychiatrie-Erfahrener e.V.', 'Selbsthilfe', 'Villingen-Schwenningen', 'kontakt-info@bpe-online.de', 'Klein & Regelmäßig'],
    ['Jugendhaus Fürstenberg', 'Kinder- & Jugendarbeit', 'Hüfingen', 'Svenja.markowitsch@huefingen.de', 'Klein & Regelmäßig'],
    ['Jugendtreff Hasen', 'Kinder- & Jugendarbeit', 'Tuningen', 'Jugendtreff-hasen@web.de', 'Klein & Regelmäßig'],
    ['Jugendclub Schabenhausen', 'Kinder- & Jugendarbeit', 'Niedereschach', 'info@jc-niederschach.de', 'Klein & Regelmäßig'],
    ['Naturparkschule Josef-Hebting-Schule', 'Natur', 'Vöhrenbach', 'sekretariat@jhs-voehrenbach.de', 'Klein & Regelmäßig'],
    ['Jugendhaus Blumberg', 'Kinder- & Jugendarbeit', 'Blumberg', 'sasa.hustic@stadt-blumberg.de', 'Klein & Regelmäßig'],
    ['Jugendhaus Behla', 'Kinder- & Jugendarbeit', 'Hüfingen', 'Svenja.markowitsch@huefingen.de', 'Klein & Regelmäßig'],
    ['Generationentreff LEBENSWert', 'Soziales', 'Bad Dürrheim', 'info@generationentreff-lebenswert.de', 'Klein & Regelmäßig'],
    ['Jugendmusikschule -- Musikalische Früherziehung', 'Musik', 'Triberg', 'jms@jugendmusikschulen.de', 'Klein & Regelmäßig'],
    ['Jugendtreff', 'Kinder- & Jugendarbeit', 'Furtwangen', 'Dmaute.stadt@furtwangen.de', 'Klein & Regelmäßig'],
    ['Reitsportgemeinschaft Gnadental', 'Sportangebot', 'Donaueschingen', '(Tel. 0771 61723)', 'Klein & Regelmäßig'],
    ['Reit- und Fahrverein Donaueschingen e.V.', 'Sportangebot', 'Donaueschingen', 'vorstand1@ruf-donaueschingen.de', 'Klein & Regelmäßig'],
    ['Jugendhaus Bohrturm', 'Kinder- & Jugendarbeit', 'Bad Dürrheim', 'jugendhaus@bad-duerrheim.de', 'Klein & Regelmäßig'],
    ['Jugendhaus Hüfingen', 'Kinder- & Jugendarbeit', 'Hüfingen', 'Svenja.markowitsch@huefingen.de', 'Klein & Regelmäßig'],
    # Cat 4
    ['Initiative Gesund leben in St. Georgen', 'Prävention', 'St. Georgen', '(Keine Mail gefunden)', 'Klein & Gelegentlich'],
    ['Blaues Kreuz in Deutschland e.V.', 'Prävention', 'St. Georgen', 'b.becker@bluprevent.de', 'Klein & Gelegentlich'],
    ['Kegel Sport Club Brigachtal', 'Sportangebot', 'Brigachtal', 'Gunnar.tr@web.de', 'Klein & Gelegentlich'],
    ['Fischereigemeinschaft Gütenbach', 'Sportangebot', 'Gütenbach', 'info@guetenbach.de', 'Klein & Gelegentlich'],
    ['Sportkegelclub Kegelfreunde Unterkirnach', 'Sportangebot', 'Unterkirnach', 'vorstand.skckegelfreundeunterkirnach@gmx.de', 'Klein & Gelegentlich'],
    ['Harmonikaverein & Josef-Hebting-Schule', 'Musik', 'Vöhrenbach', 'marinaguenter@gmx.de', 'Klein & Gelegentlich'],
    ['VHS Tuningen', 'Bildung', 'Tuningen', 'vhs@villingen-schwenningen.de', 'Klein & Gelegentlich'],
    ['Feuerwehr Schonach', 'BOS', 'Schonach', 'ts@feuerwehr-schonach.de', 'Klein & Gelegentlich'],
    ['Städtischer Familienpass', 'Kinder- & Jugendarbeit', 'Bräunlingen', 'Stephanie.engesser@braeunlingen.de', 'Klein & Gelegentlich'],
    ['SBBZ Weiherdammschule', 'Bildung / Beratung', 'Blumberg', 'info@weiherdammschule-blumberg.de', 'Klein & Gelegentlich'],
    ['Schulsozialarbeit GMS Eschach-Neckar', 'Kinder- & Jugendarbeit', 'Niedereschach', 'd_storz@lfa.org', 'Klein & Gelegentlich'],
    ['Sportkegelclub Rot-Weiß Unterkirnach', 'Sportangebot', 'Unterkirnach', 'michaelstorzukn@t-online.de', 'Klein & Gelegentlich'],
    ['Trachtenkapelle Buchenberg', 'Musik', 'Königsfeld', 'tk-buchenberg@t-online.de', 'Klein & Gelegentlich']
]

# Create Base DataFrame
cols = ['Name', 'Angebotstyp', 'Stadt', 'E-Mail', 'Kategorie']
df = pd.DataFrame(data, columns=cols)

# Enrich with Management Columns
df['Status'] = ''
df['Kontakt_Datum'] = ''
df['Nachhaken_Datum'] = ''
df['Interview_Datum'] = ''
df['Interviewer'] = 'Andreas' # Default
df['Notizen'] = ''

# --- 3. Excel Construction ---
OUTPUT_FILE = os.path.join(cleanup_dir, 'Interview_Partner_Tracking_Master.xlsx')
wb = Workbook()
if 'Sheet' in wb.sheetnames: del wb['Sheet']

# Styles
header_font = Font(bold=True, color='FFFFFF')
header_fill = PatternFill(start_color='203764', end_color='203764', fill_type='solid') # Dark Blue
sub_header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid') # Lighter Blue
center_align = Alignment(horizontal='center', vertical='center')
thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

# Dropdown Data
STATUS_LIST = ["Kontaktiert", "Nachfragen", "Zusage", "Terminiert", "Durchgeführt", "Abgeschlossen", "Absage"]
status_string = ','.join(STATUS_LIST)

# --- 3.1 COCKPIT / DASHBOARD ---
ws_dash = wb.create_sheet('Cockpit')
ws_dash.sheet_view.showGridLines = False
ws_dash.column_dimensions['A'].width = 2
ws_dash.column_dimensions['B'].width = 30
ws_dash.column_dimensions['C'].width = 15

# Title
ws_dash['B2'] = 'Projekt-Cockpit: Interviewpartner'
ws_dash['B2'].font = Font(size=20, bold=True, color='203764')

# -- Table 1: Status Overview --
row = 5
ws_dash.cell(row=row, column=2, value='Status Übersicht').font = Font(size=14, bold=True)
row += 1
ws_dash.cell(row=row, column=2, value='Status').fill = sub_header_fill
ws_dash.cell(row=row, column=2, value='Status').font = header_font
ws_dash.cell(row=row, column=3, value='Anzahl').fill = sub_header_fill
ws_dash.cell(row=row, column=3, value='Anzahl').font = header_font

start_status_row = row + 1
for st in STATUS_LIST:
    row += 1
    ws_dash.cell(row=row, column=2, value=st).border = thin_border
    # Formula pointing to Overview sheet (Table Name will be defined later or simple Range)
    # Assuming Overview data is in rows 5-100, Column F (Status is index 6)
    ws_dash.cell(row=row, column=3, value=f'=COUNTIF(Übersicht!F$5:F$200, "{st}")').border = thin_border
    ws_dash.cell(row=row, column=3).alignment = center_align

# Total
row += 1
ws_dash.cell(row=row, column=2, value='Gesamt').font = Font(bold=True)
ws_dash.cell(row=row, column=3, value=f'=SUM(C{start_status_row}:C{row-1})').font = Font(bold=True)
ws_dash.cell(row=row, column=3).alignment = center_align

# -- Table 2: Analyse nach Stadt --
row = 5
col_city = 5
ws_dash.cell(row=row, column=col_city, value='Top Städte').font = Font(size=14, bold=True)
ws_dash.column_dimensions[get_column_letter(col_city)].width = 25
row += 1
ws_dash.cell(row=row, column=col_city, value='Stadt').fill = sub_header_fill
ws_dash.cell(row=row, column=col_city).font = header_font
ws_dash.cell(row=row, column=col_city+1, value='Anzahl').fill = sub_header_fill
ws_dash.cell(row=row, column=col_city+1).font = header_font

# Top cities from data
cities = df['Stadt'].value_counts().head(8).index.tolist()
for city in cities:
    row += 1
    ws_dash.cell(row=row, column=col_city, value=city).border = thin_border
    ws_dash.cell(row=row, column=col_city+1, value=f'=COUNTIF(Übersicht!C$5:C$200, "{city}")').border = thin_border
    ws_dash.cell(row=row, column=col_city+1).alignment = center_align

# -- Table 3: Analyse nach Typ --
row = 5
col_type = 8
ws_dash.cell(row=row, column=col_type, value='Verteilung Angebotstyp').font = Font(size=14, bold=True)
ws_dash.column_dimensions[get_column_letter(col_type)].width = 30
row += 1
ws_dash.cell(row=row, column=col_type, value='Typ').fill = sub_header_fill
ws_dash.cell(row=row, column=col_type).font = header_font
ws_dash.cell(row=row, column=col_type+1, value='Anzahl').fill = sub_header_fill
ws_dash.cell(row=row, column=col_type+1).font = header_font

types = df['Angebotstyp'].value_counts().head(8).index.tolist()
for t in types:
    row += 1
    ws_dash.cell(row=row, column=col_type, value=t).border = thin_border
    ws_dash.cell(row=row, column=col_type+1, value=f'=COUNTIF(Übersicht!B$5:B$200, "{t}")').border = thin_border
    ws_dash.cell(row=row, column=col_type+1).alignment = center_align

# -- Table 4: Analyse nach Kategorie --
row = 16 # Below Status
ws_dash.cell(row=row, column=2, value='Kategorie').font = Font(size=14, bold=True)
row += 1
ws_dash.cell(row=row, column=2, value='Kategorie').fill = sub_header_fill
ws_dash.cell(row=row, column=2).font = header_font
ws_dash.cell(row=row, column=3, value='Anzahl').fill = sub_header_fill
ws_dash.cell(row=row, column=3).font = header_font

cats = df['Kategorie'].unique()
for c in cats:
    row += 1
    ws_dash.cell(row=row, column=2, value=c).border = thin_border
    ws_dash.cell(row=row, column=3, value=f'=COUNTIF(Übersicht!E$5:E$200, "{c}")').border = thin_border
    ws_dash.cell(row=row, column=3).alignment = center_align

# --- 3.2 DATA SHEETS ---
sheet_configs = {
    'Übersicht': df,
    'Groß & Regelmäßig': df[df['Kategorie'] == 'Groß & Regelmäßig'],
    'Groß & Gelegentlich': df[df['Kategorie'] == 'Groß & Gelegentlich'],
    'Klein & Regelmäßig': df[df['Kategorie'] == 'Klein & Regelmäßig'],
    'Klein & Gelegentlich': df[df['Kategorie'] == 'Klein & Gelegentlich']
}

for name, sheet_data in sheet_configs.items():
    ws = wb.create_sheet(name)
    
    # 1. Title & Count
    ws['A1'] = f'Tracking: {name}'
    ws['A1'].font = Font(size=16, bold=True, color='203764')
    
    ws['A2'] = 'Einträge:'
    ws['B2'] = f'=COUNTA(A5:A{len(sheet_data)+100})'
    ws['B2'].font = Font(bold=True, color='0000FF')
    ws['B2'].alignment = Alignment(horizontal='left')

    # 2. Headers (Row 4)
    headers = list(sheet_data.columns)
    for i, h in enumerate(headers, 1):
        cell = ws.cell(row=4, column=i, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = Border(bottom=Side(style='medium'))
        
        # Column Widths
        width = 20
        if 'Name' in h: width = 40
        elif 'Mail' in h: width = 35
        elif 'Datum' in h: width = 15
        elif 'Notizen' in h: width = 50
        ws.column_dimensions[get_column_letter(i)].width = width

    # 3. Data
    for r_idx, row_data in enumerate(sheet_data.itertuples(index=False), 5):
        for c_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            cell.border = thin_border
            
            # Center specific columns
            if c_idx > 5: # Status, Dates, etc.
                cell.alignment = Alignment(horizontal='center')

    # 4. Features
    last_row = len(sheet_data) + 100
    
    # Freeze Panes
    ws.freeze_panes = 'A5'
    
    # Auto Filter
    ws.auto_filter.ref = f'A4:{get_column_letter(len(headers))}{len(sheet_data)+4}'
    
    # Dropdown for Status (Col F / 6)
    # The header is at row 4. Data starts at 5.
    status_col_idx = headers.index('Status') + 1
    status_letter = get_column_letter(status_col_idx)
    
    dv = DataValidation(type='list', formula1=f'"{status_string}"', allow_blank=True)
    dv.error = 'Wählen Sie einen gültigen Status.'
    dv.errorTitle = 'Eingabefehler'
    ws.add_data_validation(dv)
    dv.add(f'{status_letter}5:{status_letter}{last_row}')
    
    # Conditional Formatting for Status
    # Green for "Zusage", "Terminiert", "Durchgeführt", "Abgeschlossen"
    green_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
    ws.conditional_formatting.add(f'{status_letter}5:{status_letter}{last_row}', 
                                  FormulaRule(formula=[f'OR(${status_letter}5="Zusage",${status_letter}5="Terminiert",${status_letter}5="Durchgeführt",${status_letter}5="Abgeschlossen")'], fill=green_fill))
    
    # Red for "Absage"
    red_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
    ws.conditional_formatting.add(f'{status_letter}5:{status_letter}{last_row}', 
                                  CellIsRule(operator='equal', formula=['"Absage"'], fill=red_fill))

    # Conditional Formatting for Dates (Overdue)
    # Nachhaken Datum (Col H / 8)
    if 'Nachhaken_Datum' in headers:
        nh_idx = headers.index('Nachhaken_Datum') + 1
        nh_letter = get_column_letter(nh_idx)
        # Red if date < TODAY and Status is NOT finished
        # Formula: =AND(H5<TODAY(), H5<>"", F5<>"Abgeschlossen", F5<>"Absage")
        overdue_formula = f'AND(${nh_letter}5<TODAY(), ${nh_letter}5<>"", ${status_letter}5<>"Abgeschlossen", ${status_letter}5<>"Absage")'
        ws.conditional_formatting.add(f'{nh_letter}5:{nh_letter}{last_row}',
                                      FormulaRule(formula=[overdue_formula], fill=red_fill))

# Save
wb.save(OUTPUT_FILE)
print(f'Master File Created: {OUTPUT_FILE}')
