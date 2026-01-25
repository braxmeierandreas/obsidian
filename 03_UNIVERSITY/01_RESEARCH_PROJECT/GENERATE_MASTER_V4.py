import pandas as pd
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule, FormulaRule

# --- 1. Cleanup ---
cleanup_dir = '1.Datenbank/Andreas - Analyse/Analyse-Ergebnisse/'
target_file = 'Interview_Partner_Tracking_Master.xlsx'
full_path = os.path.join(cleanup_dir, target_file)

if os.path.exists(full_path):
    try:
        os.remove(full_path)
        print(f'Deleted old version: {target_file}')
    except:
        pass

# --- 2. Configuration ---
# Updated Team
TEAM_MEMBERS = [
    "Andreas Braxmeier",
    "Nadine Beha",
    "Sophie Scheffler",
    "Ragavi Anusujantha",
    "Nina Ludwig"
]
team_string = ','.join(TEAM_MEMBERS)

STATUS_LIST = ["Kontaktiert", "Nachfragen", "Zusage", "Terminiert", "Durchgeführt", "Abgeschlossen", "Absage"]
status_string = ','.join(STATUS_LIST)

PLATFORM_LIST = ["Zoom", "Präsenz", "Telefon", "Teams", "Sonstiges"]
platform_string = ','.join(PLATFORM_LIST)

# --- 3. Data Setup (52 Partners) ---
data = [
    # Cat 1
    ['Katholisches Familienzentrum St. Konrad', 'Soziales / Selbsthilfe', 'Villingen-Schwenningen', 'kontakt@familienzentrum-stkonrad.de', 'Groß & Regelmäßig'],
    ['Ago e.V.', 'Prävention', 'Donaueschingen', 'info@verein-ago.de', 'Groß & Regelmäßig'],
    ['Psychologische Beratungsstelle Ev./Kath. Kirche', 'Beratung', 'Villingen-Schwenningen', 'gemeindebuero.schwenningen@elk-wue.de', 'Groß & Regelmäßig'],
    ['Sommerrainschule St. Georgen -- Sportangebote', 'Sportangebot', 'St. Georgen', 'sbbz-st.georgen@t-online.de', 'Groß & Regelmäßig'],
    ['Kinderfeuerwehr St. Georgen', 'BOS', 'St. Georgen', 'info@feuerwehr-st-georgen.de', 'Groß & Regelmäßig'],
    ['Neckarschule -- Klasse 2000', 'Prävention', 'Villingen-Schwenningen', 'neckarschule@villingen-schwenningen.de', 'Groß & Regelmäßig'],
    ['Georg Müller Grund- und Realschule -- Präventionskonzept', 'Prävention', 'Villingen-Schwenningen', 'info@gmsvs.de', 'Groß & Regelmäßig'],
    ['Caritas Schwangerschaftsberatung', 'Beratung', 'Villingen-Schwenningen', 'info@caritas-sbk.de', 'Groß & Regelmäßig'],
    ['Sozialpsychiatrischer Dienst (Caritas)', 'Soziales', 'Villingen-Schwenningen', 'elke.schwarz@caritas-sbk.de', 'Groß & Regelmäßig'],
    ['Malteser Hilfsdienst -- Psychosoziale Notfallvorsorge', 'BOS', 'Villingen-Schwenningen', 'Rainer.Kuehl@malteser.org', 'Groß & Regelmäßig'],
    ['AWO Haus der Kinder', 'Betreuung', 'Villingen-Schwenningen', 'hausderkinder@awo-ov-vs.de', 'Groß & Regelmäßig'],
    ['Pro Familia Villingen-Schwenningen', 'Beratung', 'Villingen-Schwenningen', 'vs-villingen@profamilia.de', 'Groß & Regelmäßig'],
    ['Sportangebot Öfingen 1969', 'Sportangebot', 'Bad Dürrheim', 'jugend@sv-oefingen.de', 'Groß & Regelmäßig'],
    # Cat 2
    ['BEKJ Villingen-Schwenningen', 'Beratung', 'Villingen-Schwenningen', 'beratungsstelle-bekj-fw@Lrasbk.de', 'Groß & Gelegentlich'],
    ['Polizei Triberg -- Präventionsvorträge', 'Prävention', 'Triberg', 'Keine E-Mail!', 'Groß & Gelegentlich'],
    ['Schwarzwald-Gymnasium', 'Beratung', 'Triberg', 'rf@schwarzwald-gymnasium.de', 'Groß & Gelegentlich'],
    ['Lernort Natur -- Jägervereinigung SBK', 'Natur', 'SBK', 'kjm-kjv-sbk@outlook.de', 'Groß & Gelegentlich'],
    ['Naturfreunde', 'Umweltschutz', 'Bräunlingen', 'naturfreunde-braeunlingen@gmx.de', 'Groß & Gelegentlich'],
    ['Kontaktstelle Beratung & Freizeit', 'Soziales', 'Donaueschingen', 'David.Mueller@Bruderhausdiakonie.de', 'Groß & Gelegentlich'],
    ['Schulnetzwerk', 'Bildung', 'St. Georgen', 'sekretariat@rsstg.de', 'Groß & Gelegentlich'],
    ['AWO Kinder- und Jugendhilfe', 'Soziales', 'Villingen-Schwenningen', 's.brunner@awo-ov-vs.de', 'Groß & Gelegentlich'],
    ['DRK -- Psychosoziale Notfallvorsorge', 'BOS', 'Villingen-Schwenningen', 'mail@psnv-sbk.de', 'Groß & Gelegentlich'],
    ['Projekt Bewegte Schule -- Dom-Clemente-Schule', 'Sportangebot', 'Schonach', 'd.tritschler@schonach.de', 'Groß & Gelegentlich'],
    ['Schönwälder Naturfreibad', 'Sportstätte', 'Schönwald', 'mail@schoenwald.de', 'Groß & Gelegentlich'],
    ['Freibad Tannheim', 'Sportstätte', 'Tannheim', 'info@freibad-tannheim.de', 'Groß & Gelegentlich'],
    ['Deutscher Kinderschutzbund OV', 'Soziales', 'Furtwangen', 'info@dksb-furtwangen.de', 'Groß & Gelegentlich'],
    ['BUND Ortsverband', 'Umweltschutz', 'Königsfeld', 'bund.sbh@bund.net', 'Groß & Gelegentlich'],
    # Cat 3
    ['Bundesverband Psychiatrie-Erfahrener e.V.', 'Selbsthilfe', 'Villingen-Schwenningen', 'kontakt-info@bpe-online.de', 'Klein & Regelmäßig'],
    ['Jugendhaus Fürstenberg', 'Kinder- & Jugendarbeit', 'Hüfingen', 'Svenja.markowitsch@huefingen.de', 'Klein & Regelmäßig'],
    ['Jugendtreff Hasen', 'Kinder- & Jugendarbeit', 'Tuningen', 'Jugendtreff-hasen@web.de', 'Klein & Regelmäßig'],
    ['Jugendclub Schabenhausen', 'Kinder- & Jugendarbeit', 'Niedereschach', 'info@jc-niederschach.de', 'Klein & Regelmäßig'],
    ['Naturparkschule Josef-Hebting-Schule', 'Natur', 'Vöhrenbach', 'sekretariat@jhs-voehrenbach.de', 'Klein & Regelmäßig'],
    ['Jugendhaus Blumberg', 'Kinder- & Jugendarbeit', 'Blumberg', 'sasa.hustic@stadt-blumberg.de', 'Klein & Regelmäßig'],
    ['Jugendhaus Behla', 'Kinder- & Jugendarbeit', 'Hüfingen', 'Svenja.markowitsch@huefingen.de', 'Klein & Regelmäßig'],
    ['Generationentreff LEBENSWert', 'Soziales', 'Bad Dürrheim', 'info@generationentreff-lebenswert.de', 'Klein & Regelmäßig'],
    ['Jugendmusikschule -- Musikalische Früherziehung', 'Musik', 'Triberg', 'jms@jugendmusikschulen.de', 'Klein & Regelmäßig'],
    ['Jugendtreff', 'Kinder- & Jugendarbeit', 'Furtwangen', 'Dmaute.stadt@furtwangen.de', 'Klein & Regelmäßig'],
    ['Reitsportgemeinschaft Gnadental', 'Sportangebot', 'Donaueschingen', '(Tel. 0771 61723)', 'Klein & Regelmäßig'],
    ['Reit- und Fahrverein Donaueschingen e.V.', 'Sportangebot', 'Donaueschingen', 'vorstand1@ruf-donaueschingen.de', 'Klein & Regelmäßig'],
    ['Jugendhaus Bohrturm', 'Kinder- & Jugendarbeit', 'Bad Dürrheim', 'jugendhaus@bad-duerrheim.de', 'Klein & Regelmäßig'],
    ['Jugendhaus Hüfingen', 'Kinder- & Jugendarbeit', 'Hüfingen', 'Svenja.markowitsch@huefingen.de', 'Klein & Regelmäßig'],
    # Cat 4
    ['Initiative Gesund leben in St. Georgen', 'Prävention', 'St. Georgen', '(Keine Mail gefunden)', 'Klein & Gelegentlich'],
    ['Blaues Kreuz in Deutschland e.V.', 'Prävention', 'St. Georgen', 'b.becker@bluprevent.de', 'Klein & Gelegentlich'],
    ['Kegel Sport Club Brigachtal', 'Sportangebot', 'Brigachtal', 'Gunnar.tr@web.de', 'Klein & Gelegentlich'],
    ['Fischereigemeinschaft Gütenbach', 'Sportangebot', 'Gütenbach', 'info@guetenbach.de', 'Klein & Gelegentlich'],
    ['Sportkegelclub Kegelfreunde Unterkirnach', 'Sportangebot', 'Unterkirnach', 'vorstand.skckegelfreundeunterkirnach@gmx.de', 'Klein & Gelegentlich'],
    ['Harmonikaverein & Josef-Hebting-Schule', 'Musik', 'Vöhrenbach', 'marinaguenter@gmx.de', 'Klein & Gelegentlich'],
    ['VHS Tuningen', 'Bildung', 'Tuningen', 'vhs@villingen-schwenningen.de', 'Klein & Gelegentlich'],
    ['Feuerwehr Schonach', 'BOS', 'Schonach', 'ts@feuerwehr-schonach.de', 'Klein & Gelegentlich'],
    ['Städtischer Familienpass', 'Kinder- & Jugendarbeit', 'Bräunlingen', 'Stephanie.engesser@braeunlingen.de', 'Klein & Gelegentlich'],
    ['SBBZ Weiherdammschule', 'Bildung / Beratung', 'Blumberg', 'info@weiherdammschule-blumberg.de', 'Klein & Gelegentlich'],
    ['Schulsozialarbeit GMS Eschach-Neckar', 'Kinder- & Jugendarbeit', 'Niedereschach', 'd_storz@lfa.org', 'Klein & Gelegentlich'],
    ['Sportkegelclub Rot-Weiß Unterkirnach', 'Sportangebot', 'Unterkirnach', 'michaelstorzukn@t-online.de', 'Klein & Gelegentlich'],
    ['Trachtenkapelle Buchenberg', 'Musik', 'Königsfeld', 'tk-buchenberg@t-online.de', 'Klein & Gelegentlich']
]

# Create Dataframe
cols = ['Name', 'Angebotstyp', 'Stadt', 'E-Mail', 'Kategorie']
df = pd.DataFrame(data, columns=cols)

# Enrich with Logic Columns
df['Status'] = ''
df['Ansprechpartner'] = ''
df['Interviewer 1'] = ''
df['Interviewer 2'] = ''
df['Datum'] = ''
df['Uhrzeit'] = ''
df['Dauer'] = ''
df['Art'] = '' # Platform
df['Ort / Link'] = ''
df['Themen'] = ''
df['Unterlagen'] = ''

# --- 4. Excel Generation ---
wb = Workbook()
if 'Sheet' in wb.sheetnames: del wb['Sheet']

# Styles
header_font = Font(bold=True, color='FFFFFF')
header_fill = PatternFill(start_color='203764', end_color='203764', fill_type='solid') 
sub_header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
center_align = Alignment(horizontal='center', vertical='center')
thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

# -----------------
# SHEET: Cockpit
# -----------------
ws_dash = wb.create_sheet('Cockpit')
ws_dash.sheet_view.showGridLines = False
ws_dash.column_dimensions['A'].width = 2
ws_dash.column_dimensions['B'].width = 30
ws_dash.column_dimensions['C'].width = 15

ws_dash['B2'] = 'Projekt-Cockpit: Interviewplanung'
ws_dash['B2'].font = Font(size=20, bold=True, color='203764')

# -- Table 1: Status Overview (Col F in Overview) --
row = 5
ws_dash.cell(row=row, column=2, value='Status Übersicht').font = Font(size=14, bold=True)
row += 1
ws_dash.cell(row=row, column=2, value='Status').fill = sub_header_fill
ws_dash.cell(row=row, column=2).font = header_font
ws_dash.cell(row=row, column=3, value='Anzahl').fill = sub_header_fill
ws_dash.cell(row=row, column=3).font = header_font

for st in STATUS_LIST:
    row += 1
    ws_dash.cell(row=row, column=2, value=st).border = thin_border
    ws_dash.cell(row=row, column=3, value=f'=COUNTIF(Übersicht!F$5:F$500, "{st}")').border = thin_border
    ws_dash.cell(row=row, column=3).alignment = center_align

# -- Table 2: Team Workload (Columns H and I for Interviewer 1 & 2) --
row_wl = 5
col_wl = 5
ws_dash.cell(row=row_wl, column=col_wl, value='Team Auslastung').font = Font(size=14, bold=True)
ws_dash.column_dimensions[get_column_letter(col_wl)].width = 25
row_wl += 1
ws_dash.cell(row=row_wl, column=col_wl, value='Name').fill = sub_header_fill
ws_dash.cell(row=row_wl, column=col_wl).font = header_font
ws_dash.cell(row=row_wl, column=col_wl+1, value='Einsätze').fill = sub_header_fill
ws_dash.cell(row=row_wl, column=col_wl+1).font = header_font

for member in TEAM_MEMBERS:
    row_wl += 1
    ws_dash.cell(row=row_wl, column=col_wl, value=member).border = thin_border
    # Interviewer 1 = H (8), Interviewer 2 = I (9)
    formula = f'=COUNTIF(Übersicht!H$5:H$500, "{member}") + COUNTIF(Übersicht!I$5:I$500, "{member}")'
    ws_dash.cell(row=row_wl, column=col_wl+1, value=formula).border = thin_border
    ws_dash.cell(row=row_wl, column=col_wl+1).alignment = center_align

# -- Table 3: Cities (Col C in Overview) --
row_city = 5
col_city = 8
ws_dash.cell(row=row_city, column=col_city, value='Top Städte').font = Font(size=14, bold=True)
ws_dash.column_dimensions[get_column_letter(col_city)].width = 25
row_city += 1
ws_dash.cell(row=row_city, column=col_city, value='Stadt').fill = sub_header_fill
ws_dash.cell(row=row_city, column=col_city).font = header_font
ws_dash.cell(row=row_city, column=col_city+1, value='Anzahl').fill = sub_header_fill
ws_dash.cell(row=row_city, column=col_city+1).font = header_font

# Get top cities dynamically from data
top_cities = df['Stadt'].value_counts().head(7).index.tolist()
for c in top_cities:
    row_city += 1
    ws_dash.cell(row=row_city, column=col_city, value=c).border = thin_border
    ws_dash.cell(row=row_city, column=col_city+1, value=f'=COUNTIF(Übersicht!C$5:C$500, "{c}")').border = thin_border
    ws_dash.cell(row=row_city, column=col_city+1).alignment = center_align

# -- Table 4: Types (Col B in Overview) --
row_type = 5
col_type = 11
ws_dash.cell(row=row_type, column=col_type, value='Top Angebotstypen').font = Font(size=14, bold=True)
ws_dash.column_dimensions[get_column_letter(col_type)].width = 35
row_type += 1
ws_dash.cell(row=row_type, column=col_type, value='Typ').fill = sub_header_fill
ws_dash.cell(row=row_type, column=col_type).font = header_font
ws_dash.cell(row=row_type, column=col_type+1, value='Anzahl').fill = sub_header_fill
ws_dash.cell(row=row_type, column=col_type+1).font = header_font

top_types = df['Angebotstyp'].value_counts().head(7).index.tolist()
for t in top_types:
    row_type += 1
    ws_dash.cell(row=row_type, column=col_type, value=t).border = thin_border
    ws_dash.cell(row=row_type, column=col_type+1, value=f'=COUNTIF(Übersicht!B$5:B$500, "{t}")').border = thin_border
    ws_dash.cell(row=row_type, column=col_type+1).alignment = center_align

# -----------------
# SHEET: Infos & Ressourcen
# -----------------
ws_info = wb.create_sheet('Infos & Ressourcen')
ws_info['A1'] = 'Wichtige Projekt-Informationen'
ws_info['A1'].font = Font(size=18, bold=True, color='203764')

info_headers = ['Thema', 'Link / Ort', 'Benutzername', 'Passwort / Notiz']
for i, h in enumerate(info_headers, 1):
    c = ws_info.cell(row=3, column=i, value=h)
    c.font = header_font
    c.fill = header_fill
    c.alignment = center_align

info_data = [
    ['Zoom Meeting Raum', '', '', ''],
    ['Projektordner (Drive)', '', '', ''],
    ['Miro Board', '', '', ''],
    ['Interviewleitfaden (Link)', '', '', '']
]
for r_idx, row_data in enumerate(info_data, 4):
    for c_idx, val in enumerate(row_data, 1):
        ws_info.cell(row=r_idx, column=c_idx, value=val).border = thin_border

ws_info.column_dimensions['A'].width = 30
ws_info.column_dimensions['B'].width = 50
ws_info.column_dimensions['C'].width = 25
ws_info.column_dimensions['D'].width = 30

# -----------------
# DATA SHEETS
# -----------------
sheets = {
    'Übersicht': df,
    'Groß & Regelmäßig': df[df['Kategorie'] == 'Groß & Regelmäßig'],
    'Groß & Gelegentlich': df[df['Kategorie'] == 'Groß & Gelegentlich'],
    'Klein & Regelmäßig': df[df['Kategorie'] == 'Klein & Regelmäßig'],
    'Klein & Gelegentlich': df[df['Kategorie'] == 'Klein & Gelegentlich']
}

for name, data in sheets.items():
    ws = wb.create_sheet(name)
    
    # Header
    ws['A1'] = name.upper()
    ws['A1'].font = Font(size=14, bold=True)
    ws['B1'] = f'Anzahl: {len(data)}'
    
    # Table Headers
    cols = list(data.columns)
    for i, col in enumerate(cols, 1):
        c = ws.cell(row=4, column=i, value=col)
        c.font = header_font
        c.fill = header_fill
        c.alignment = center_align
        c.border = Border(bottom=Side(style='medium'))
        
        # Widths
        width = 20
        if 'Name' in col: width = 40
        elif 'Mail' in col: width = 35
        elif 'Ort' in col or 'Link' in col: width = 30
        elif 'Themen' in col: width = 40
        elif 'Interviewer' in col: width = 25
        ws.column_dimensions[get_column_letter(i)].width = width

    # Data
    for r_idx, row in enumerate(data.itertuples(index=False), 5):
        for c_idx, val in enumerate(row, 1):
            ws.cell(row=r_idx, column=c_idx, value=str(val) if val else '').border = thin_border

    # Dropdowns & Validation
    last_row = len(data) + 100
    
    # 1. Status (Col F / 6)
    dv_status = DataValidation(type="list", formula1=f'"{status_string}"', allow_blank=True)
    ws.add_data_validation(dv_status)
    dv_status.add(f'F5:F{last_row}')
    
    # 2. Team (Interviewer 1=H/8, Interviewer 2=I/9)
    dv_team = DataValidation(type="list", formula1=f'"{team_string}"', allow_blank=True)
    ws.add_data_validation(dv_team)
    dv_team.add(f'H5:I{last_row}')
    
    # 3. Art/Platform (Col M / 13)
    try:
        art_idx = cols.index('Art') + 1
        art_letter = get_column_letter(art_idx)
        dv_art = DataValidation(type="list", formula1=f'"{platform_string}"', allow_blank=True)
        ws.add_data_validation(dv_art)
        dv_art.add(f'{art_letter}5:{art_letter}{last_row}')
    except: pass

    # Layout
    ws.freeze_panes = 'A5'
    ws.auto_filter.ref = f'A4:{get_column_letter(len(cols))}{len(data)+4}'
    
    # Conditional Formatting
    green = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
    red = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
    
    ws.conditional_formatting.add(f'F5:F{last_row}', FormulaRule(formula=['OR(F5="Zusage",F5="Terminiert",F5="Durchgeführt")'], fill=green))
    ws.conditional_formatting.add(f'F5:F{last_row}', CellIsRule(operator='equal', formula=['"Absage"'], fill=red))

wb.save(full_path)
print(f'Master File Saved: {full_path}')
