import pandas as pd
import re
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# --- Data ---
data = [
    # Cat 1
    ['Katholisches Familienzentrum St. Konrad', 'Soziales / Selbsthilfe', 'Villingen-Schwenningen', 'kontakt@familienzentrum-stkonrad.de', 'Groß & Regelmäßig'],
    ['Ago e.V.', 'Prävention', 'Donaueschingen', 'info@verein-ago.de', 'Groß & Regelmäßig'],
    ['Psychologische Beratungsstelle Ev./Kath. Kirche', 'Beratung', 'Villingen-Schwenningen', 'gemeindebuero.schwenningen@elk-wue.de', 'Groß & Regelmäßig'],
    ['Sommerrainschule St. Georgen -- Sportangebote', 'Sportangebot', 'St. Georgen', 'sbbz-st.georgen@t-online.de', 'Groß & Regelmäßig'],
    ['Kinderfeuerwehr St. Georgen', 'BOS', 'St. Georgen', 'info@feuerwehr-st-georgen.de', 'Groß & Regelmäßig'],
    ['Neckarschule -- Klasse 2000', 'Prävention', 'Villingen-Schwenningen', 'neckarschule@villingen-schwenningen.de', 'Groß & Regelmäßig'],
    ['Georg Müller Grund- und Realschule -- Präventionskonzept', 'Prävention', 'Villingen-Schwenningen', 'info@gmsvs.de', 'Groß & Regelmäßig'],
    ['Caritas Schwangerschaftsberatung', 'Beratung', 'Villingen-Schwenningen', 'info@caritas-sbk.de', 'Groß & Regelmäßig'],
    ['Sozialpsychiatrischer Dienst (Caritas)', 'Soziales', 'Villingen-Schwenningen', 'elke.schwarz@caritas-sbk.de', 'Groß & Regelmäßig'],
    ['Malteser Hilfsdienst -- Psychosoziale Notfallvorsorge', 'BOS', 'Villingen-Schwenningen', 'Rainer.Kuehl@malteser.org', 'Groß & Regelmäßig'],
    ['AWO Haus der Kinder', 'Betreuung', 'Villingen-Schwenningen', 'hausderkinder@awo-ov-vs.de', 'Groß & Regelmäßig'],
    ['Pro Familia Villingen-Schwenningen', 'Beratung', 'Villingen-Schwenningen', 'vs-villingen@profamilia.de', 'Groß & Regelmäßig'],
    ['Sportangebot Öfingen 1969', 'Sportangebot', 'Bad Dürrheim', 'jugend@sv-oefingen.de', 'Groß & Regelmäßig'],
    # Cat 2
    ['BEKJ Villingen-Schwenningen', 'Beratung', 'Villingen-Schwenningen', 'beratungsstelle-bekj-fw@Lrasbk.de', 'Groß & Gelegentlich'],
    ['Polizei Triberg -- Präventionsvorträge', 'Prävention', 'Triberg', 'Keine E-Mail!', 'Groß & Gelegentlich'],
    ['Schwarzwald-Gymnasium', 'Beratung', 'Triberg', 'rf@schwarzwald-gymnasium.de', 'Groß & Gelegentlich'],
    ['Lernort Natur -- Jägervereinigung SBK', 'Natur', 'SBK', 'kjm-kjv-sbk@outlook.de', 'Groß & Gelegentlich'],
    ['Naturfreunde', 'Umweltschutz', 'Bräunlingen', 'naturfreunde-braeunlingen@gmx.de', 'Groß & Gelegentlich'],
    ['Kontaktstelle Beratung & Freizeit', 'Soziales', 'Donaueschingen', 'David.Mueller@Bruderhausdiakonie.de', 'Groß & Gelegentlich'],
    ['Schulnetzwerk', 'Bildung', 'St. Georgen', 'sekretariat@rsstg.de', 'Groß & Gelegentlich'],
    ['AWO Kinder- und Jugendhilfe', 'Soziales', 'Villingen-Schwenningen', 's.brunner@awo-ov-vs.de', 'Groß & Gelegentlich'],
    ['DRK -- Psychosoziale Notfallvorsorge', 'BOS', 'Villingen-Schwenningen', 'mail@psnv-sbk.de', 'Groß & Gelegentlich'],
    ['Projekt Bewegte Schule -- Dom-Clemente-Schule', 'Sportangebot', 'Schonach', 'd.tritschler@schonach.de', 'Groß & Gelegentlich'],
    ['Schönwälder Naturfreibad', 'Sportstätte', 'Schönwald', 'mail@schoenwald.de', 'Groß & Gelegentlich'],
    ['Freibad Tannheim', 'Sportstätte', 'Tannheim', 'info@freibad-tannheim.de', 'Groß & Gelegentlich'],
    ['Deutscher Kinderschutzbund OV', 'Soziales', 'Furtwangen', 'info@dksb-furtwangen.de', 'Groß & Gelegentlich'],
    ['BUND Ortsverband', 'Umweltschutz', 'Königsfeld', 'bund.sbh@bund.net', 'Groß & Gelegentlich'],
    # Cat 3
    ['Bundesverband Psychiatrie-Erfahrener e.V.', 'Selbsthilfe', 'Villingen-Schwenningen', 'kontakt-info@bpe-online.de', 'Klein & Regelmäßig'],
    ['Jugendhaus Fürstenberg', 'Kinder- & Jugendarbeit', 'Hüfingen', 'Svenja.markowitsch@huefingen.de', 'Klein & Regelmäßig'],
    ['Jugendtreff Hasen', 'Kinder- & Jugendarbeit', 'Tuningen', 'Jugendtreff-hasen@web.de', 'Klein & Regelmäßig'],
    ['Jugendclub Schabenhausen', 'Kinder- & Jugendarbeit', 'Niedereschach', 'info@jc-niederschach.de', 'Klein & Regelmäßig'],
    ['Naturparkschule Josef-Hebting-Schule', 'Natur', 'Vöhrenbach', 'sekretariat@jhs-voehrenbach.de', 'Klein & Regelmäßig'],
    ['Jugendhaus Blumberg', 'Kinder- & Jugendarbeit', 'Blumberg', 'sasa.hustic@stadt-blumberg.de', 'Klein & Regelmäßig'],
    ['Jugendhaus Behla', 'Kinder- & Jugendarbeit', 'Hüfingen', 'Svenja.markowitsch@huefingen.de', 'Klein & Regelmäßig'],
    ['Generationentreff LEBENSWert', 'Soziales', 'Bad Dürrheim', 'info@generationentreff-lebenswert.de', 'Klein & Regelmäßig'],
    ['Jugendmusikschule -- Musikalische Früherziehung', 'Musik', 'Triberg', 'jms@jugendmusikschulen.de', 'Klein & Regelmäßig'],
    ['Jugendtreff', 'Kinder- & Jugendarbeit', 'Furtwangen', 'Dmaute.stadt@furtwangen.de', 'Klein & Regelmäßig'],
    ['Reitsportgemeinschaft Gnadental', 'Sportangebot', 'Donaueschingen', '(Tel. 0771 61723)', 'Klein & Regelmäßig'],
    ['Reit- und Fahrverein Donaueschingen e.V.', 'Sportangebot', 'Donaueschingen', 'vorstand1@ruf-donaueschingen.de', 'Klein & Regelmäßig'],
    ['Jugendhaus Bohrturm', 'Kinder- & Jugendarbeit', 'Bad Dürrheim', 'jugendhaus@bad-duerrheim.de', 'Klein & Regelmäßig'],
    ['Jugendhaus Hüfingen', 'Kinder- & Jugendarbeit', 'Hüfingen', 'Svenja.markowitsch@huefingen.de', 'Klein & Regelmäßig'],
    # Cat 4
    ['Initiative Gesund leben in St. Georgen', 'Prävention', 'St. Georgen', '(Keine Mail gefunden)', 'Klein & Gelegentlich'],
    ['Blaues Kreuz in Deutschland e.V.', 'Prävention', 'St. Georgen', 'b.becker@bluprevent.de', 'Klein & Gelegentlich'],
    ['Kegel Sport Club Brigachtal', 'Sportangebot', 'Brigachtal', 'Gunnar.tr@web.de', 'Klein & Gelegentlich'],
    ['Fischereigemeinschaft Gütenbach', 'Sportangebot', 'Gütenbach', 'info@guetenbach.de', 'Klein & Gelegentlich'],
    ['Sportkegelclub Kegelfreunde Unterkirnach', 'Sportangebot', 'Unterkirnach', 'vorstand.skckegelfreundeunterkirnach@gmx.de', 'Klein & Gelegentlich'],
    ['Harmonikaverein & Josef-Hebting-Schule', 'Musik', 'Vöhrenbach', 'marinaguenter@gmx.de', 'Klein & Gelegentlich'],
    ['VHS Tuningen', 'Bildung', 'Tuningen', 'vhs@villingen-schwenningen.de', 'Klein & Gelegentlich'],
    ['Feuerwehr Schonach', 'BOS', 'Schonach', 'ts@feuerwehr-schonach.de', 'Klein & Gelegentlich'],
    ['Städtischer Familienpass', 'Kinder- & Jugendarbeit', 'Bräunlingen', 'Stephanie.engesser@braeunlingen.de', 'Klein & Gelegentlich'],
    ['SBBZ Weiherdammschule', 'Bildung / Beratung', 'Blumberg', 'info@weiherdammschule-blumberg.de', 'Klein & Gelegentlich'],
    ['Schulsozialarbeit GMS Eschach-Neckar', 'Kinder- & Jugendarbeit', 'Niedereschach', 'd_storz@lfa.org', 'Klein & Gelegentlich'],
    ['Sportkegelclub Rot-Weiß Unterkirnach', 'Sportangebot', 'Unterkirnach', 'michaelstorzukn@t-online.de', 'Klein & Gelegentlich'],
    ['Trachtenkapelle Buchenberg', 'Musik', 'Königsfeld', 'tk-buchenberg@t-online.de', 'Klein & Gelegentlich']
]

columns = ['Name', 'Angebotstyp', 'Stadt/Gemeinde', 'E-Mail', 'Kategorie']
df_partners = pd.DataFrame(data, columns=columns)
# Add Status Column
df_partners['Status'] = '' 

# --- Excel Config ---
OUTPUT_FILE = '1.Datenbank/Andreas - Analyse/Analyse-Ergebnisse/Interview_Partner_Tracking_Final.xlsx'
wb = Workbook()
if 'Sheet' in wb.sheetnames: del wb['Sheet']

sheets = {
    'Übersicht': df_partners,
    'Groß & Regelmäßig': df_partners[df_partners['Kategorie'] == 'Groß & Regelmäßig'],
    'Groß & Gelegentlich': df_partners[df_partners['Kategorie'] == 'Groß & Gelegentlich'],
    'Klein & Regelmäßig': df_partners[df_partners['Kategorie'] == 'Klein & Regelmäßig'],
    'Klein & Gelegentlich': df_partners[df_partners['Kategorie'] == 'Klein & Gelegentlich']
}

STATUS_OPTIONS = [
    "Kontaktiert", "Nachfragen", "Zusage", "Terminiert", "Durchgeführt", "Abgeschlossen", "Absage"
]
STATUS_STRING = ','.join(STATUS_OPTIONS)

for name, data in sheets.items():
    ws = wb.create_sheet(name)
    
    # --- 1. Dashboard ---
    ws['A1'] = 'Status Dashboard'
    ws['A1'].font = Font(size=14, bold=True)
    
    ws['A3'] = 'Status'
    ws['B3'] = 'Anzahl'
    ws['A3'].font = Font(bold=True, color='FFFFFF')
    ws['B3'].font = Font(bold=True, color='FFFFFF')
    ws['A3'].fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    ws['B3'].fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    
    start_dashboard_row = 4
    for idx, status in enumerate(STATUS_OPTIONS):
        r = start_dashboard_row + idx
        ws.cell(row=r, column=1, value=status)
        ws.cell(row=r, column=2, value=f'=COUNTIF($F$15:$F$500, "{status}")')
        
    total_row = start_dashboard_row + len(STATUS_OPTIONS)
    ws.cell(row=total_row, column=1, value='Gesamt')
    ws.cell(row=total_row, column=2, value=f'=SUM(B{start_dashboard_row}:B{total_row-1})')
    ws.cell(row=total_row, column=1).font = Font(bold=True)
    ws.cell(row=total_row, column=2).font = Font(bold=True)
    
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    for r in range(3, total_row + 1):
        for c in range(1, 3):
            ws.cell(row=r, column=c).border = thin_border

    # --- 2. Data Table ---
    data_start_row = 14
    header_row = data_start_row
    
    cols = list(data.columns)
    for i, col in enumerate(cols, 1):
        c = ws.cell(row=header_row, column=i, value=str(col))
        c.font = Font(bold=True, color='FFFFFF')
        c.fill = PatternFill(start_color='ED7D31', end_color='ED7D31', fill_type='solid')
        c.alignment = Alignment(horizontal='center')
        
    for r_idx, row in enumerate(data.itertuples(index=False), header_row + 1):
        for c_idx, val in enumerate(row, 1):
            ws.cell(row=r_idx, column=c_idx, value=str(val) if val else '')
            
    # --- 3. Dropdown ---
    dv = DataValidation(type="list", formula1=f'"{STATUS_STRING}"', allow_blank=True)
    dv.error = 'Bitte wählen Sie einen Wert aus der Liste.'
    dv.errorTitle = 'Ungültiger Status'
    ws.add_data_validation(dv)
    
    status_col_letter = get_column_letter(len(cols))
    dv.add(f'{status_col_letter}{header_row+1}:{status_col_letter}{header_row+100}')

    # --- 4. Layout ---
    ws.freeze_panes = f'A{header_row+1}'
    ws.auto_filter.ref = f'A{header_row}:{get_column_letter(len(cols))}{len(data)+header_row}'
    
    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 25
    ws.column_dimensions['D'].width = 35
    ws.column_dimensions['E'].width = 25
    ws.column_dimensions['F'].width = 20

wb.save(OUTPUT_FILE)
print(f'Successfully created {OUTPUT_FILE}')
